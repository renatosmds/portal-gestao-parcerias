import csv
import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils import timezone

from apps.empresas.models import Empresa
from apps.termos.models import Termos
from apps.prestacao.models import Prestacao
from apps.lancamentos.models import Lancamento


def normalizar_chave(valor):
    texto = str(valor or "").strip().lower()
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def ler_arquivo(arquivo):
    nome = arquivo.name.lower()
    if nome.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Instale openpyxl para importar planilhas XLSX.") from exc
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active
        valores = list(ws.iter_rows(values_only=True))
        if not valores:
            return [], []
        cab = [normalizar_chave(v) for v in valores[0]]
        linhas = [dict(zip(cab, ["" if v is None else v for v in row])) for row in valores[1:] if any(v not in (None, "") for v in row)]
        return cab, linhas

    bruto = arquivo.read()
    texto = bruto.decode("utf-8-sig", errors="replace")
    amostra = texto[:4096]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    leitor = csv.DictReader(io.StringIO(texto), dialect=dialect)
    cab = [normalizar_chave(c) for c in (leitor.fieldnames or [])]
    linhas = []
    for original in leitor:
        linha = {normalizar_chave(k): (v or "").strip() for k, v in original.items()}
        if any(str(v).strip() for v in linha.values()):
            linhas.append(linha)
    return cab, linhas


def valor(linha, *nomes):
    for nome in nomes:
        v = linha.get(normalizar_chave(nome))
        if v not in (None, ""):
            return str(v).strip()
    return ""


def decimal_br(v):
    if v in (None, ""):
        return None
    txt = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return Decimal(txt)
    except InvalidOperation:
        raise ValueError(f"Valor monetário inválido: {v}")


def data_iso(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    txt = str(v).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(txt, formato).date()
        except ValueError:
            continue
    raise ValueError(f"Data inválida: {v}")


def validar_linhas(tipo, linhas):
    erros = []
    for i, linha in enumerate(linhas, start=2):
        try:
            if tipo == "osc" and not valor(linha, "nome", "razao_social", "osc"):
                raise ValueError("Nome/razão social é obrigatório")
            if tipo == "termo" and not valor(linha, "numtermo", "numero_termo", "termo"):
                raise ValueError("Número do Termo é obrigatório")
            if tipo == "prestacao" and not valor(linha, "numtermo", "numero_termo"):
                raise ValueError("Número do Termo é obrigatório")
            if tipo == "lancamento":
                if not valor(linha, "empresa", "osc", "nomeosc"):
                    raise ValueError("Empresa/OSC é obrigatória")
                if not valor(linha, "numero_lancamento", "lancamento", "numero"):
                    raise ValueError("Número do lançamento é obrigatório")
                decimal_br(valor(linha, "valor", "valor_documento"))
                data_iso(valor(linha, "data_documento", "data"))
        except ValueError as exc:
            erros.append({"linha": i, "erro": str(exc), "dados": linha})
    return erros


@transaction.atomic
def confirmar_importacao(importacao):
    novos = atualizados = duplicados = erros_count = 0
    erros = []
    for i, linha in enumerate(importacao.linhas, start=2):
        try:
            if importacao.tipo == "osc":
                nome = valor(linha, "nome", "razao_social", "osc")
                _, criado = Empresa.objects.get_or_create(nome__iexact=nome, defaults={"nome": nome})
                novos += int(criado); duplicados += int(not criado)

            elif importacao.tipo == "termo":
                numero = valor(linha, "numtermo", "numero_termo", "termo")
                defaults = {
                    "nomeosc": valor(linha, "nomeosc", "osc", "empresa")[:50] or None,
                    "tipo": valor(linha, "tipo", "modalidade")[:100] or None,
                    "objeto": valor(linha, "objeto")[:100] or None,
                    "inicioVigencia": valor(linha, "inicio_vigencia", "inicio")[:100] or None,
                    "terminoVigencia": valor(linha, "termino_vigencia", "fim")[:100] or None,
                    "valorglobal": decimal_br(valor(linha, "valor_global", "valorglobal")),
                    "valorrepasse": decimal_br(valor(linha, "valor_repasse", "valorrepasse")),
                    "status": valor(linha, "status", "situacao")[:100] or None,
                    "numpa": valor(linha, "processo_sei", "numpa", "processo")[:10] or None,
                }
                _, criado = Termos.objects.update_or_create(numtermo=numero, defaults=defaults)
                novos += int(criado); atualizados += int(not criado)

            elif importacao.tipo == "prestacao":
                numero = valor(linha, "numtermo", "numero_termo")
                tipo_termo = valor(linha, "tipo_termo", "tipo") or "TC"
                defaults = {
                    "credor": valor(linha, "credor", "osc", "empresa")[:50] or None,
                    "CpfCnpj": valor(linha, "cpf_cnpj", "cnpj")[:18] or None,
                    "valorContrato": float(decimal_br(valor(linha, "valor_contrato", "valor")) or 0),
                    "situacao_workflow": valor(linha, "situacao", "status") or Prestacao.SituacaoWorkflow.ELABORACAO,
                }
                _, criado = Prestacao.objects.update_or_create(numtermo=numero, tipoTermo=tipo_termo, defaults=defaults)
                novos += int(criado); atualizados += int(not criado)

            elif importacao.tipo == "lancamento":
                nome_empresa = valor(linha, "empresa", "osc", "nomeosc")
                empresa, _ = Empresa.objects.get_or_create(nome__iexact=nome_empresa, defaults={"nome": nome_empresa})
                numero = valor(linha, "numero_lancamento", "lancamento", "numero")
                termo_num = valor(linha, "numtermo", "numero_termo", "termo")
                termo = Termos.objects.filter(numtermo=termo_num).first() if termo_num else None
                defaults = {
                    "termo": termo,
                    "tipo_documento": valor(linha, "tipo_documento", "tipo") or Lancamento.TipoDocumento.OUTRO,
                    "numero_documento": valor(linha, "numero_documento", "documento")[:80],
                    "data_documento": data_iso(valor(linha, "data_documento", "data")),
                    "data_pagamento": data_iso(valor(linha, "data_pagamento")),
                    "descricao": valor(linha, "descricao", "historico")[:255],
                    "valor_documento": decimal_br(valor(linha, "valor", "valor_documento")) or Decimal("0"),
                    "criado_por": importacao.criado_por,
                }
                _, criado = Lancamento.objects.update_or_create(empresa=empresa, numero_lancamento=numero, defaults=defaults)
                novos += int(criado); atualizados += int(not criado)
        except Exception as exc:
            erros_count += 1
            erros.append({"linha": i, "erro": str(exc), "dados": linha})

    importacao.total_novos = novos
    importacao.total_atualizados = atualizados
    importacao.total_duplicados = duplicados
    importacao.total_erros = erros_count
    importacao.erros = erros
    importacao.confirmado_em = timezone.now()
    importacao.situacao = "parcial" if erros_count else "confirmada"
    importacao.save()
    return importacao
