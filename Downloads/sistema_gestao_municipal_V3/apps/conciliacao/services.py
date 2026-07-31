import csv
import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction

from .models import ImportacaoExtrato, Movimentacao, OcorrenciaConciliacao


def _decimal(valor):
    if valor in (None, ""):
        return None
    if isinstance(valor, Decimal):
        return valor
    txt = str(valor).strip().replace("R$", "").replace(" ", "")
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return Decimal(txt)
    except InvalidOperation as exc:
        raise ValueError(f"Valor inválido: {valor}") from exc


def _data(valor):
    if hasattr(valor, "date"):
        return valor.date()
    if hasattr(valor, "year") and not isinstance(valor, str):
        return valor
    txt = str(valor or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y%m%d"):
        try:
            return datetime.strptime(txt[:10], fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Data inválida: {valor}")


def _normalizar(valor):
    return re.sub(r"[^a-z0-9]+", "_", str(valor or "").lower()).strip("_")


def _linha_para_mov(linha):
    n = {_normalizar(k): v for k, v in linha.items()}
    def pegar(*chaves):
        for chave in chaves:
            valor = n.get(_normalizar(chave))
            if valor not in (None, ""):
                return valor
        return ""
    credito = _decimal(pegar("credito", "valor_credito"))
    debito = _decimal(pegar("debito", "valor_debito"))
    valor_generico = _decimal(pegar("valor", "amount"))
    tipo_txt = str(pegar("tipo", "natureza", "trntype")).lower()
    if credito is not None:
        tipo, valor = Movimentacao.Tipo.CREDITO, abs(credito)
    elif debito is not None:
        tipo, valor = Movimentacao.Tipo.DEBITO, abs(debito)
    elif valor_generico is not None:
        tipo = Movimentacao.Tipo.DEBITO if valor_generico < 0 or tipo_txt in {"d", "debito", "debit"} else Movimentacao.Tipo.CREDITO
        valor = abs(valor_generico)
    else:
        raise ValueError("Informe crédito, débito ou valor.")
    descricao = str(pegar("descricao", "historico", "memo", "name") or "Movimentação bancária").strip()
    return {
        "data": _data(pegar("data", "data_movimento", "dtposted")),
        "descricao": descricao[:255],
        "documento": str(pegar("documento", "numero_documento", "checknum"))[:80],
        "favorecido": str(pegar("favorecido", "beneficiario", "payee"))[:180],
        "tipo": tipo,
        "valor": valor,
        "saldo_apos": _decimal(pegar("saldo", "saldo_apos", "balance")),
    }


def _ler_csv(arquivo):
    texto = arquivo.read().decode("utf-8-sig", errors="replace")
    try:
        dialeto = csv.Sniffer().sniff(texto[:4096], delimiters=";,	,")
    except csv.Error:
        dialeto = csv.excel
        dialeto.delimiter = ";"
    return list(csv.DictReader(io.StringIO(texto), dialect=dialeto))


def _ler_xlsx(arquivo):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Instale openpyxl para importar XLSX.") from exc
    ws = load_workbook(arquivo, read_only=True, data_only=True).active
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []
    cab = [str(v or "") for v in linhas[0]]
    return [dict(zip(cab, row)) for row in linhas[1:] if any(v not in (None, "") for v in row)]


def _ler_ofx(arquivo):
    texto = arquivo.read().decode("latin-1", errors="replace")
    blocos = re.findall(r"<STMTTRN>(.*?)(?:</STMTTRN>|<STMTTRN>)", texto, flags=re.I | re.S)
    linhas = []
    for bloco in blocos:
        def tag(nome):
            m = re.search(rf"<{nome}>([^<\r\n]+)", bloco, flags=re.I)
            return m.group(1).strip() if m else ""
        linhas.append({"dtposted": tag("DTPOSTED")[:8], "amount": tag("TRNAMT"), "memo": tag("MEMO") or tag("NAME"), "checknum": tag("CHECKNUM"), "trntype": tag("TRNTYPE")})
    return linhas


@transaction.atomic
def importar_extrato(conciliacao, arquivo, usuario=None):
    ext = arquivo.name.lower().rsplit(".", 1)[-1]
    leitores = {"csv": _ler_csv, "xlsx": _ler_xlsx, "ofx": _ler_ofx}
    linhas = leitores[ext](arquivo)
    arquivo.seek(0)
    importacao = ImportacaoExtrato.objects.create(conciliacao=conciliacao, arquivo=arquivo, formato=ext.upper(), total_linhas=len(linhas), criado_por=usuario)
    importadas, erros = 0, []
    for numero, linha in enumerate(linhas, start=2):
        try:
            dados = _linha_para_mov(linha)
            _, criada = Movimentacao.objects.get_or_create(conciliacao=conciliacao, importacao=importacao, **dados)
            importadas += int(criada)
        except Exception as exc:
            erros.append({"linha": numero, "erro": str(exc)})
    importacao.total_importadas = importadas
    importacao.total_erros = len(erros)
    importacao.erros = erros
    importacao.situacao = ImportacaoExtrato.Situacao.COM_ERROS if erros else ImportacaoExtrato.Situacao.PROCESSADA
    importacao.save(update_fields=["total_importadas", "total_erros", "erros", "situacao"])
    gerar_ocorrencias(conciliacao)
    conciliacao.recalcular_situacao()
    return importacao


def gerar_ocorrencias(conciliacao):
    OcorrenciaConciliacao.objects.filter(conciliacao=conciliacao, situacao=OcorrenciaConciliacao.Situacao.PENDENTE).delete()
    for mov in conciliacao.movimentacoes.filter(situacao=Movimentacao.Situacao.PENDENTE):
        OcorrenciaConciliacao.objects.create(conciliacao=conciliacao, movimentacao=mov, tipo=OcorrenciaConciliacao.Tipo.MOVIMENTO_SEM_LANCAMENTO, descricao=f"Movimentação de {mov.valor} ainda não vinculada a lançamento.")
    for mov in conciliacao.movimentacoes.filter(categoria=Movimentacao.Categoria.TARIFA):
        OcorrenciaConciliacao.objects.get_or_create(conciliacao=conciliacao, movimentacao=mov, tipo=OcorrenciaConciliacao.Tipo.TARIFA, defaults={"descricao": "Tarifa bancária identificada; verifique autorização e tratamento."})
    if conciliacao.saldo_final_informado is not None and abs(conciliacao.diferenca or Decimal("0.00")) > Decimal("0.01"):
        OcorrenciaConciliacao.objects.create(conciliacao=conciliacao, tipo=OcorrenciaConciliacao.Tipo.SALDO, descricao=f"Diferença entre saldo informado e calculado: {conciliacao.diferenca}.")
